from typing import Any
import openpyxl
from datetime import datetime, timedelta
from front import app, reset, sum_work_hours


def calculate_working_hours_from_excel(filename, sheet_name):
    DAILY: datetime = timedelta(hours=8, minutes=30)
    workbook: openpyxl.Workbook = openpyxl.load_workbook(filename)
    sheet: Any = workbook[sheet_name]

    # הגדרת טווח התאים (תשנה בהתאם למבנה הקובץ שלך)
    start_row = 2  # שורה התחלתית של הנתונים
    end_row = 5  # שורה סופית של הנתונים
    start_col = 1  # עמודה התחלתית של הנתונים
    end_col = 16  # עמודה סופית של הנתונים

    total_hours: dict = {}
    spare_hours: dict = {}
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col, 3):
            _date = sheet.cell(row=row, column=col).value
            in_time = sheet.cell(row=row, column=col + 1).value
            out_time = sheet.cell(row=row, column=col + 2).value
            if in_time and out_time:
                in_time = datetime.strptime(str(in_time), "%H:%M:%S")
                out_time = datetime.strptime(str(out_time), "%H:%M:%S")
                total_hours[_date.strftime("%d/%m")] = ":".join(
                    str(out_time - in_time).split(":")[:2]
                )
                spare_hours[_date.strftime("%d/%m")] = ":".join(
                    str(out_time - in_time - DAILY).split(":")[:2]
                )
            else:
                total_hours[_date.strftime("%d/%m")] = "0:00"
                spare_hours[_date.strftime("%d/%m")] = "0:00"

    return total_hours, spare_hours


def calculate_total_time(time_list):
    total_seconds = 0
    for time_str in time_list:
        # המרת מחרוזת ל-datetime.timedelta
        time_delta = datetime.strptime(time_str, "%H:%M:%S") - datetime(1900, 1, 1)
        total_seconds += time_delta.total_seconds()

    # חישוב שעות ודקות
    total_hours, remainder = divmod(total_seconds, 3600)
    total_minutes = remainder // 60

    return int(total_hours), int(total_minutes)


if __name__ == "__main__":
    filename = "h.xlsx"
    sheet_name = "Dec"
    total, spare = calculate_working_hours_from_excel(filename, sheet_name)

    print(sum_work_hours(spare))
    reset(total, spare)

    app.run_server(debug=True, port=7457)
